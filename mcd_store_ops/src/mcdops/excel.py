"""把整个 package 导出成一本 Excel 工作簿（11 个 sheet）。

设计原则和 HTML 看板一样，但多一条 Excel 特有的：

    **凡是能算的都写公式，不写死值。**

所以每张表分两类列：
  - 浅蓝底 = 原子输入（净销售、交易笔数、工时、废弃金额…），店长可以直接改
  - 白底   = 公式（客单价、达成率、成本率、SPLH、状态灯…），跟着输入自动重算

阈值只存在一个地方——「目标与口径」这张表。其余所有 sheet 的红黄绿都用
INDEX/MATCH 回查它，所以把食品成本率的绿线从 31% 改成 30%，整本工作簿的
颜色和优先级排序会一起变。这是 Excel 版相对 HTML 版真正多出来的价值。

仅用 Excel 2007 级函数（INDEX/MATCH/IFERROR/COUNTIF/SUMPRODUCT/TEXT），
不用 XLOOKUP / FILTER / SORT / UNIQUE —— 那些在多数环境下无法求值。
"""

from __future__ import annotations

from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from . import matrix as mx, report as rp, schema
from .metrics import MetricEngine, Slice, severity

# ── 视觉常量（对齐常见 QSR 管理模板的配色） ───────────────────────────────
FONT = "Arial"
BRAND = "B3261E"      # 标题条 深红
BRAND_INK = "7A1F1A"  # 副标题文字
CREAM = "FFF8E7"      # 副标题底
NAVY = "203040"       # 表头底 / 正文字色
INPUT_BG = "EAF3F8"   # 浅蓝 = 可编辑输入
BAND = "F7F9FB"       # 斑马纹
RED, AMBER, GREEN = "FDECEC", "FFF3CD", "E8F5E9"
RED_INK, AMBER_INK, GREEN_INK = "B3261E", "8A6100", "1B5E20"

THIN = Side(style="thin", color="D8DEE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_MONEY = '#,##0'
FMT_MONEY2 = '#,##0.00'
FMT_PCT = '0.0%'
FMT_PCT2 = '0.00%'
FMT_NUM1 = '0.0'
FMT_NUM2 = '0.00'
FMT_INT = '#,##0'
FMT_DATE = 'yyyy-mm-dd'

TARGETS = "目标与口径"      # 单一事实来源，所有阈值回查它
DAILY = "每日经营日报"

# 「目标与口径」表的行区间（写完后回填）
T_FIRST, T_LAST = 5, 5


# ── 小工具 ───────────────────────────────────────────────────────────────
def _title_block(ws, title: str, subtitle: str, ncols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, title)
    c.font = Font(name=FONT, size=16, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=BRAND)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, subtitle)
    c.font = Font(name=FONT, size=9, color=BRAND_INK)
    c.fill = PatternFill("solid", fgColor=CREAM)
    c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[2].height = 28
    ws.sheet_view.showGridLines = False


def _header(ws, row: int, headers: list[str], widths: list[int] | None = None) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, h)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 30
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def _cell(ws, row: int, col: int, value, fmt: str | None = None, *,
          is_input: bool = False, bold: bool = False, align: str | None = None,
          wrap: bool = False, band: bool = False, size: float = 9):
    c = ws.cell(row, col, value)
    c.font = Font(name=FONT, size=size, bold=bold, color=NAVY)
    if fmt:
        c.number_format = fmt
    if is_input:
        c.fill = PatternFill("solid", fgColor=INPUT_BG)
    elif band:
        c.fill = PatternFill("solid", fgColor=BAND)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = BORDER
    return c


def _note(ws, row: int, text: str, ncols: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font = Font(name=FONT, size=8.5, italic=True, color="6B7280")
    c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 26


def _section(ws, row: int, text: str, ncols: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font = Font(name=FONT, size=10.5, bold=True, color=NAVY)
    c.fill = PatternFill("solid", fgColor="EEF2F6")
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def _status_cf(ws, ref: str) -> None:
    """给写着 绿/黄/红 的列加条件格式。"""
    for text, bg, ink in (("红", RED, RED_INK), ("黄", AMBER, AMBER_INK), ("绿", GREEN, GREEN_INK)):
        ws.conditional_formatting.add(ref, CellIsRule(
            operator="equal", formula=[f'"{text}"'],
            fill=PatternFill("solid", bgColor=bg), font=Font(name=FONT, size=9, bold=True, color=ink)))


def _add_table(ws, name: str, ref: str) -> None:
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
    ws.add_table(t)


# 指标阈值回查：都指向「目标与口径」，改那一张表就能带动全本
def _lookup(col: str, metric_id: str) -> str:
    return (f"INDEX('{TARGETS}'!${col}${T_FIRST}:${col}${T_LAST},"
            f"MATCH(\"{metric_id}\",'{TARGETS}'!$A${T_FIRST}:$A${T_LAST},0))")


def _status_formula(value_ref: str, metric_id: str) -> str:
    """按方向和绿黄线判红黄绿。方向也从「目标与口径」读，所以改方向也生效。"""
    green, amber, direction = _lookup("G", metric_id), _lookup("H", metric_id), _lookup("E", metric_id)
    higher = f'{direction}="越高越好"'
    up = f'IF({value_ref}>={green},"绿",IF({value_ref}>={amber},"黄","红"))'
    down = f'IF({value_ref}<={green},"绿",IF({value_ref}<={amber},"黄","红"))'
    return f'=IFERROR(IF({higher},{up},{down}),"—")'


def _severity_formula(value_ref: str, metric_id: str) -> str:
    """超线几个黄灯带宽。见 docs/03 第 3.4 节。"""
    green, amber, direction = _lookup("G", metric_id), _lookup("H", metric_id), _lookup("E", metric_id)
    dist = f'IF({direction}="越高越好",{green}-{value_ref},{value_ref}-{green})'
    return f'=IFERROR(MAX(0,{dist}/ABS({amber}-{green})),0)'


# ── 单位换算：引擎里 % 是 31.2，Excel 里要存 0.312 ─────────────────────────
def _to_sheet(value, unit: str):
    if value is None:
        return None
    return value / 100.0 if unit == "%" else value


def _fmt_for(unit: str) -> str:
    return {"%": FMT_PCT, "CNY": FMT_MONEY, "笔": FMT_INT, "秒": FMT_INT,
            "分": FMT_NUM1, "次": FMT_INT, "分钟": FMT_INT,
            "CNY/工时": FMT_MONEY, "件": FMT_NUM2, "件/千单": FMT_NUM2}.get(unit, FMT_NUM2)


DIRECTION_CN = {"higher_better": "越高越好", "lower_better": "越低越好", "neutral": "越低越好（上限）"}


# ══ 每日原子数据 ══════════════════════════════════════════════════════════
def _atoms(engine: MetricEngine, d: date) -> dict:
    """一天的原子量。派生指标在 Excel 里用公式算，这里只取分子分母。"""
    s = Slice(engine.data, d, engine.profile, None, None)
    txns = s.day_rows("fct_pos_transaction")
    kds = s.day_rows("fct_kds_order")
    dt = s.day_rows("fct_dt_timer")
    deliv = s.day_rows("fct_delivery_order")
    waste = s.day_rows("fct_waste_log")
    batches = s.day_rows("fct_holding_batch")
    sched = s.day_rows("fct_schedule")
    cards = s.day_rows("fct_timecard")
    surveys = [r for r in s.day_rows("fct_guest_survey") if r["source"] == "receipt"]
    temps = s.day_rows("fct_temperature_log")
    walks = s.day_rows("fct_travel_path_check")
    fc_day = s.day_rows("agg_food_cost_day")
    lab_day = s.day_rows("agg_labor_day")

    net = sum(t["net_amount"] for t in txns)
    gc = len(txns)
    front = [k for k in kds if k["channel"] in ("FrontCounter", "Kiosk", "MobileApp")]

    def avg(xs):
        xs = list(xs)
        return sum(xs) / len(xs) if xs else 0

    return {
        "forecast_sales": sum(f["forecast_sales"] for f in s.day_rows("dim_sales_forecast")),
        "net_sales": net,
        "guest_count": gc,
        "digital_share": (sum(1 for t in txns if t["channel"] in ("Kiosk", "MobileApp", "Delivery"))
                          / gc if gc else 0),
        "delivery_sales": sum(t["net_amount"] for t in txns if t["channel"] == "Delivery"),
        "item_qty": sum(l["qty"] for l in s.day_rows("fct_pos_line_item")),
        "food_cost": fc_day[0]["actual_cost"] if fc_day else 0,
        "waste_cost": sum(w["cost"] for w in waste),
        "labor_hours": lab_day[0]["actual_hours"] if lab_day else 0,
        "planned_hours": lab_day[0]["planned_hours"] if lab_day else 0,
        # 班组工时（不含管理层的固定工时）：排班偏差的正确分母
        "crew_planned": sum(x["planned_crew"] for x in s.day_rows("agg_labor_interval")) * 0.25,
        "crew_actual": sum(x["actual_crew"] for x in s.day_rows("agg_labor_interval")) * 0.25,
        "wage": engine.profile["labor"]["avg_wage_per_hour"],
        "dt_total": avg(x["total_seconds"] for x in dt),
        "dt_window": avg(x["window_seconds"] for x in dt),
        "kds_prep": avg(k["prep_seconds"] for k in kds),
        "front_otd": avg(k["service_seconds"] for k in front),
        "delivery_handover": avg(x["handover_seconds"] for x in deliv),
        "on_time_orders": sum(1 for k in kds if k["on_time"]),
        "kds_orders": len(kds),
        "issues": len(s.day_rows("fct_order_issue")),
        "fs_score": engine.compute(d, ["food_safety_score"])["food_safety_score"] or 0,
        "temp_fail": sum(1 for r in temps if not r["in_range"]),
        "batch_made": sum(b["qty_made"] for b in batches),
        "batch_discard": sum(b["qty_discard"] for b in batches),
        "clean_score": avg(w["score"] for w in walks),
        "downtime": sum(t["downtime_minutes"] for t in s.day_rows("fct_equipment_ticket")),
        "sched_n": len(sched),
        "card_n": len(cards),
        "late_n": sum(1 for c in cards if c["late_minutes"] > 5),
        "cert_n": sum(1 for x in sched if x["certified"]),
        "turnover": engine.compute(d, ["crew_turnover_30d"])["crew_turnover_30d"] or 0,
        "survey_sum": sum(r["score"] for r in surveys),
        "survey_n": len(surveys),
        "delivery_rating": avg(x["rating"] for x in deliv),
        "cash_var": sum(abs(r["variance"]) for r in s.day_rows("fct_cash_declaration")),
        "upsell_manual": sum(1 for t in txns if t["cashier_id"] != "KIOSK"),
        "upsell_hit": sum(1 for t in txns if t["cashier_id"] != "KIOSK" and t["has_upsell"]),
    }


# ══ Sheet 1 · 使用说明 ════════════════════════════════════════════════════
def _sheet_readme(wb, engine, biz_date, days) -> None:
    ws = wb.create_sheet("使用说明")
    p = engine.profile
    N = 8
    _title_block(ws, "门店运营报表包 · 使用说明",
                 "本工作簿是快餐门店运营的通用模拟模板，不代表任何特定品牌的真实数据或内部标准；"
                 "所有阈值均为示例，须按各自门店预算、合同与当地法规替换。", N)
    for i, w in enumerate([16, 26, 34, 26, 12, 26, 14, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 4
    _section(ws, r, "0. 怎么用这本工作簿", N); r += 1
    for line in [
        ("① 先改「目标与口径」", "那是全本唯一的阈值来源。改一个绿线，看板、优先级矩阵、"
                                "所有归因表的红黄绿会一起变。"),
        ("② 每天填「每日经营日报」", "只填浅蓝色格子（原子数）。白色格子是公式，会自己算，别覆盖。"),
        ("③ 早会看「店长看板」", "四个数 + 六盏灯 + 今天的三件事。看完就去干活，别在这里停留超过 1 分钟。"),
        ("④ 要找原因才翻矩阵", "时段矩阵找“哪半小时崩了”，交叉矩阵找“哪个渠道 × 哪个时段”，"
                               "归因矩阵找“哪一格拖了全店最多”。"),
        ("⑤ 「报表清单」是节奏表", "谁在什么时间看哪张表、看完做什么。新店长交接先看它。"),
    ]:
        _cell(ws, r, 1, line[0], bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=N)
        _cell(ws, r, 2, line[1], wrap=True)
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    _section(ws, r, "1. 颜色与格式约定", N); r += 1
    _header(ws, r, ["标记", "含义", "说明", "", "", "", "", ""]); r += 1
    for mark, mean, desc, fill in [
        ("浅蓝底", "可编辑输入", "原子数：净销售、交易笔数、工时、废弃金额…… 每天由店长/系统填入", INPUT_BG),
        ("白底", "公式", "派生指标：客单价、达成率、成本率、SPLH、状态灯…… 自动计算，不要手改", None),
        ("绿", "达标", "在绿线以内", GREEN),
        ("黄", "关注", "越过绿线但没到红线；单日出现是噪声，连续 3 天才是信号", AMBER),
        ("红", "告警", "越过红线；当天必须有动作或有上报", RED),
    ]:
        c = _cell(ws, r, 1, mark, align="center", bold=True)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        _cell(ws, r, 2, mean)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=N)
        _cell(ws, r, 3, desc, wrap=True)
        r += 1

    r += 1
    _section(ws, r, "2. 一天的管理节奏（详见「报表清单」）", N); r += 1
    _header(ws, r, ["时间点", "看什么", "重点问题", "输出", "时长", "数据来源", "负责人", "完成"])
    r += 1
    first = r
    for rep in rp.DAILY_REPORTS:
        _cell(ws, r, 1, rep["when"])
        _cell(ws, r, 2, rep["name"], wrap=True)
        _cell(ws, r, 3, rep["question"], wrap=True)
        _cell(ws, r, 4, rep["action"].split("；")[0][:38], wrap=True)
        _cell(ws, r, 5, f"{rep['minutes']} 分钟", align="center")
        _cell(ws, r, 6, "、".join(rep["tables"][:2]), wrap=True)
        _cell(ws, r, 7, mx.OWNER_LABEL[rep["owner"]], align="center")
        _cell(ws, r, 8, "☐", align="center", is_input=True)
        ws.row_dimensions[r].height = 30
        r += 1
    _cell(ws, r, 1, "合计", bold=True)
    _cell(ws, r, 5, f"=SUM(E{first}:E{r-1})&\" 分钟\"", align="center", bold=True)
    ws.cell(r, 5).value = f'=TEXT(SUMPRODUCT(--SUBSTITUTE(E{first}:E{r-1}," 分钟","")),"0")&" 分钟/天"'
    r += 2

    _note(ws, r, f"门店：{p['store_name']}（{p['store_id']}，{p['format']}） · "
                 f"数据区间：{biz_date - timedelta(days=days-1)} 至 {biz_date} · "
                 f"生成命令：python -m mcdops.cli excel", N)


# ══ Sheet 2 · 目标与口径（单一事实来源） ══════════════════════════════════
def _sheet_targets(wb, engine) -> None:
    global T_FIRST, T_LAST
    ws = wb.create_sheet(TARGETS)
    N = 14
    _title_block(ws, "目标与口径 · Targets, Thresholds & KPI Definitions",
                 "全本唯一的阈值来源。浅蓝格子可改；改完「店长看板」「优先级矩阵」"
                 "「归因矩阵」的红黄绿与排序会一起变。示例阈值须按各自门店标准替换。", N)
    headers = ["指标ID", "指标", "家族", "定义 / 公式", "方向", "目标", "绿线", "黄线",
               "单位", "责任人", "频率", "影响度", "可控度", "行动卡"]
    widths = [22, 18, 12, 44, 14, 10, 10, 10, 10, 12, 8, 9, 9, 9]
    _header(ws, 4, headers, widths)

    order = {"SALES": 0, "SPEED": 1, "COST": 2, "QSC": 3, "PEOPLE": 4, "GUEST": 5}
    metrics = sorted(engine.defs.values(), key=lambda m: (order[m["family"]], m["id"]))

    r = 5
    T_FIRST = r
    for m in metrics:
        unit = m["unit"]
        fmt = FMT_PCT if unit == "%" else _fmt_for(unit)
        band = (r % 2 == 0)
        _cell(ws, r, 1, m["id"], band=band)
        _cell(ws, r, 2, m["name_cn"], bold=True, band=band)
        _cell(ws, r, 3, mx.FAMILY_LABEL[m["family"]], align="center", band=band)
        _cell(ws, r, 4, m["formula"], wrap=True, band=band)
        _cell(ws, r, 5, DIRECTION_CN[m["direction"]], align="center", band=band)
        for col, key in ((6, "target"), (7, "green"), (8, "amber")):
            _cell(ws, r, col, _to_sheet(m[key], unit), fmt, is_input=True, align="right")
        _cell(ws, r, 9, unit, align="center", band=band)
        _cell(ws, r, 10, mx.OWNER_LABEL[m["owner"]], align="center", band=band)
        _cell(ws, r, 11, mx.CADENCE_LABEL[m["cadence"]], align="center", band=band)
        _cell(ws, r, 12, m["impact"], FMT_NUM2, is_input=True, align="center")
        _cell(ws, r, 13, m["controllability"], FMT_NUM2, is_input=True, align="center")
        _cell(ws, r, 14, m["playbook"], align="center", band=band)
        ws.row_dimensions[r].height = 26
        r += 1
    T_LAST = r - 1

    ws.freeze_panes = "C5"
    r += 1
    _note(ws, r, "影响度 = 这个指标动一动对生意和顾客的影响有多大；可控度 = 店长今天真的能改变它多少。"
                 "两者都是门店层面的常识判断，不是算出来的——原材料涨价推高的食品成本率，"
                 "影响度高但可控度低，硬压只会逼出偷工减料。优先级矩阵用它们分四象限。", N)
    r += 1
    _note(ws, r, "口径说明：百分比一律以分数存储（0.312 显示为 31.2%）。"
                 "废弃率、食品成本率、千单投诉数采用「全店分母」口径（可加，用于归因）；"
                 "人工成本率、SPLH、速度类采用「切片分母」口径（各块自己的体检值）。详见 docs/03。", N)


# ══ Sheet 3 · 每日经营日报 ════════════════════════════════════════════════
# (列名, 宽, 类型)  类型: in=原子输入 / fx=公式 / id=日期
DAILY_COLS = [
    ("日期", 11, "id"), ("星期", 6, "fx"),
    ("销售预测", 11, "in"), ("净销售额", 11, "in"), ("交易笔数", 10, "in"),
    ("客单价", 9, "fx"), ("销售达成率", 11, "fx"),
    ("商品件数", 10, "in"), ("单均件数", 9, "fx"),
    ("数字渠道占比", 12, "in"), ("外送销售额", 11, "in"), ("外送占比", 10, "fx"),
    ("人工点单数", 11, "in"), ("附加销售单数", 12, "in"), ("附加销售率", 11, "fx"),
    ("食品成本额", 11, "in"), ("食品成本率", 11, "fx"),
    ("废弃金额", 10, "in"), ("废弃率", 9, "fx"),
    ("实际工时", 10, "in"), ("计划工时", 10, "in"), ("平均时薪", 10, "in"),
    ("班组计划工时", 12, "in"), ("班组实际工时", 12, "in"),
    ("人工成本", 10, "fx"), ("人工成本率", 11, "fx"), ("人时销售SPLH", 12, "fx"),
    ("工时排班偏差", 12, "fx"),
    ("得来速总时长", 12, "in"), ("得来速取餐窗", 12, "in"), ("后厨制作时长", 12, "in"),
    ("前台交付OTD", 12, "in"), ("外送出餐时长", 12, "in"),
    ("达标订单数", 11, "in"), ("服务速度达标率", 13, "fx"),
    ("差错事件数", 11, "in"), ("订单准确率", 11, "fx"), ("千单投诉数", 11, "fx"),
    ("食安日检得分", 12, "in"), ("关键温度异常", 12, "in"),
    ("保温批次数", 11, "in"), ("保温废弃数", 11, "in"), ("保温超时率", 11, "fx"),
    ("清洁巡检得分", 12, "in"), ("设备停机分钟", 12, "in"),
    ("排班人次", 10, "in"), ("到岗人次", 10, "in"), ("出勤达成率", 11, "fx"),
    ("迟到人次", 10, "in"), ("迟到率", 9, "fx"),
    ("认证人次", 10, "in"), ("岗位认证覆盖率", 13, "fx"),
    ("30日流失率", 11, "in"),
    ("调研总分", 10, "in"), ("调研份数", 10, "in"), ("顾客满意度", 11, "fx"),
    ("外卖平台评分", 12, "in"), ("现金短溢", 10, "in"),
    ("店长备注", 22, "in"),
]

# 列字母 → 便于写公式
DC = {name: get_column_letter(i) for i, (name, _w, _t) in enumerate(DAILY_COLS, start=1)}


def _sheet_daily(wb, engine, biz_date, days) -> list[date]:
    ws = wb.create_sheet(DAILY)
    N = len(DAILY_COLS)
    _title_block(ws, "每日经营日报 · Daily Operations Report",
                 "一行一天。浅蓝色是每天要填的原子数，白色是公式（客单价、各类比率、"
                 "SPLH 等）自动计算。改任何一个浅蓝格子，右边的派生列和「店长看板」会一起更新。", N)
    _header(ws, 4, [c[0] for c in DAILY_COLS], [c[1] for c in DAILY_COLS])

    dates = [biz_date - timedelta(days=i) for i in range(days - 1, -1, -1)]
    r = 5
    for d in dates:
        a = _atoms(engine, d)
        R = r  # 行号，公式里用

        def put(name, value, fmt=None, is_input=True, align=None):
            col = list(DC).index(name) + 1
            return _cell(ws, R, col, value, fmt, is_input=is_input, align=align)

        def fx(name, formula, fmt=None):
            col = list(DC).index(name) + 1
            return _cell(ws, R, col, formula, fmt, is_input=False, align="right")

        put("日期", d, FMT_DATE)
        fx("星期", f'=TEXT({DC["日期"]}{R},"aaa")')
        put("销售预测", round(a["forecast_sales"], 2), FMT_MONEY)
        put("净销售额", round(a["net_sales"], 2), FMT_MONEY)
        put("交易笔数", a["guest_count"], FMT_INT)
        fx("客单价", f'=IFERROR({DC["净销售额"]}{R}/{DC["交易笔数"]}{R},0)', FMT_MONEY2)
        fx("销售达成率", f'=IFERROR({DC["净销售额"]}{R}/{DC["销售预测"]}{R},0)', FMT_PCT)
        put("商品件数", a["item_qty"], FMT_INT)
        fx("单均件数", f'=IFERROR({DC["商品件数"]}{R}/{DC["交易笔数"]}{R},0)', FMT_NUM2)
        put("数字渠道占比", a["digital_share"], FMT_PCT)
        put("外送销售额", round(a["delivery_sales"], 2), FMT_MONEY)
        fx("外送占比", f'=IFERROR({DC["外送销售额"]}{R}/{DC["净销售额"]}{R},0)', FMT_PCT)
        put("人工点单数", a["upsell_manual"], FMT_INT)
        put("附加销售单数", a["upsell_hit"], FMT_INT)
        fx("附加销售率", f'=IFERROR({DC["附加销售单数"]}{R}/{DC["人工点单数"]}{R},0)', FMT_PCT)
        put("食品成本额", round(a["food_cost"], 2), FMT_MONEY)
        fx("食品成本率", f'=IFERROR({DC["食品成本额"]}{R}/{DC["净销售额"]}{R},0)', FMT_PCT)
        put("废弃金额", round(a["waste_cost"], 2), FMT_MONEY)
        fx("废弃率", f'=IFERROR({DC["废弃金额"]}{R}/{DC["净销售额"]}{R},0)', FMT_PCT2)
        put("实际工时", round(a["labor_hours"], 2), FMT_NUM1)
        put("计划工时", round(a["planned_hours"], 2), FMT_NUM1)
        put("平均时薪", a["wage"], FMT_MONEY2)
        put("班组计划工时", round(a["crew_planned"], 4), FMT_NUM1)
        put("班组实际工时", round(a["crew_actual"], 4), FMT_NUM1)
        fx("人工成本", f'={DC["实际工时"]}{R}*{DC["平均时薪"]}{R}', FMT_MONEY)
        fx("人工成本率", f'=IFERROR({DC["人工成本"]}{R}/{DC["净销售额"]}{R},0)', FMT_PCT)
        fx("人时销售SPLH", f'=IFERROR({DC["净销售额"]}{R}/{DC["实际工时"]}{R},0)', FMT_MONEY)
        # 只看班组：管理层工时是固定量，混进来会把偏差率稀释掉
        fx("工时排班偏差",
           f'=IFERROR(ABS({DC["班组实际工时"]}{R}-{DC["班组计划工时"]}{R})'
           f'/{DC["班组计划工时"]}{R},0)', FMT_PCT)
        put("得来速总时长", round(a["dt_total"], 2), FMT_INT)
        put("得来速取餐窗", round(a["dt_window"], 2), FMT_INT)
        put("后厨制作时长", round(a["kds_prep"], 2), FMT_INT)
        put("前台交付OTD", round(a["front_otd"], 2), FMT_INT)
        put("外送出餐时长", round(a["delivery_handover"], 2), FMT_INT)
        put("达标订单数", a["on_time_orders"], FMT_INT)
        fx("服务速度达标率", f'=IFERROR({DC["达标订单数"]}{R}/{DC["交易笔数"]}{R},0)', FMT_PCT)
        put("差错事件数", a["issues"], FMT_INT)
        fx("订单准确率", f'=IFERROR(1-{DC["差错事件数"]}{R}/{DC["交易笔数"]}{R},0)', FMT_PCT2)
        fx("千单投诉数", f'=IFERROR({DC["差错事件数"]}{R}/{DC["交易笔数"]}{R}*1000,0)', FMT_NUM2)
        put("食安日检得分", round(a["fs_score"], 3), FMT_NUM1)
        put("关键温度异常", a["temp_fail"], FMT_INT)
        put("保温批次数", a["batch_made"], FMT_INT)
        put("保温废弃数", a["batch_discard"], FMT_INT)
        fx("保温超时率", f'=IFERROR({DC["保温废弃数"]}{R}/{DC["保温批次数"]}{R},0)', FMT_PCT)
        put("清洁巡检得分", round(a["clean_score"], 3), FMT_NUM1)
        put("设备停机分钟", a["downtime"], FMT_INT)
        put("排班人次", a["sched_n"], FMT_INT)
        put("到岗人次", a["card_n"], FMT_INT)
        fx("出勤达成率", f'=IFERROR({DC["到岗人次"]}{R}/{DC["排班人次"]}{R},0)', FMT_PCT)
        put("迟到人次", a["late_n"], FMT_INT)
        fx("迟到率", f'=IFERROR({DC["迟到人次"]}{R}/{DC["到岗人次"]}{R},0)', FMT_PCT)
        put("认证人次", a["cert_n"], FMT_INT)
        fx("岗位认证覆盖率", f'=IFERROR({DC["认证人次"]}{R}/{DC["排班人次"]}{R},0)', FMT_PCT)
        put("30日流失率", a["turnover"] / 100.0, FMT_PCT)
        put("调研总分", a["survey_sum"], FMT_INT)
        put("调研份数", a["survey_n"], FMT_INT)
        fx("顾客满意度", f'=IFERROR({DC["调研总分"]}{R}/{DC["调研份数"]}{R}*20,0)', FMT_NUM1)
        put("外卖平台评分", round(a["delivery_rating"], 4), FMT_NUM2)
        put("现金短溢", round(a["cash_var"], 2), FMT_MONEY2)
        put("店长备注", "")
        r += 1

    last = r - 1
    ws.freeze_panes = "C5"
    _add_table(ws, "DailyOpsTable", f"A4:{get_column_letter(N)}{last}")

    # 条件格式：达成率色阶 + 几个关键率的红线
    ws.conditional_formatting.add(
        f'{DC["销售达成率"]}5:{DC["销售达成率"]}{last}',
        ColorScaleRule(start_type="num", start_value=0.9, start_color=RED,
                       mid_type="num", mid_value=0.98, mid_color=AMBER,
                       end_type="num", end_value=1.05, end_color=GREEN))
    for col, op, val in ((DC["服务速度达标率"], "lessThan", "0.8"),
                         (DC["订单准确率"], "lessThan", "0.99"),
                         (DC["出勤达成率"], "lessThan", "0.92")):
        ws.conditional_formatting.add(f'{col}5:{col}{last}', CellIsRule(
            operator=op, formula=[val], fill=PatternFill("solid", bgColor=RED),
            font=Font(name=FONT, size=9, color=RED_INK)))
    for col, val in ((DC["食品成本率"], "0.325"), (DC["废弃率"], "0.015"),
                     (DC["人工成本率"], "0.155"), (DC["迟到率"], "0.08")):
        ws.conditional_formatting.add(f'{col}5:{col}{last}', CellIsRule(
            operator="greaterThan", formula=[val], fill=PatternFill("solid", bgColor=RED),
            font=Font(name=FONT, size=9, color=RED_INK)))

    _note(ws, last + 2, "浅蓝 = 每天填的原子数（来自 POS / KDS / 得来速计时 / 排班考勤 / 废弃登记 / "
                        "食安检查 / 顾客调研）。白色 = 公式，勿改。"
                        "条件格式里的红线为示例硬编码值，与「目标与口径」保持一致时才准确——"
                        "改了阈值请同步这里的条件格式，或直接看「店长看板」的状态列（那里是回查公式，永远准）。", N)
    return dates


# ══ Sheet 4 · 店长看板 ════════════════════════════════════════════════════
KPI_ROWS = [
    ("sales_vs_forecast", "销售达成率"), ("net_sales", "净销售额"),
    ("guest_count", "交易笔数"), ("avg_check", "客单价"),
    ("food_cost_pct", "食品成本率"), ("waste_pct", "废弃率"),
    ("labor_cost_pct", "人工成本率"), ("splh", "人时销售 SPLH"),
    ("speed_hit_rate", "服务速度达标率"), ("dt_total_time", "得来速总时长"),
    ("order_accuracy", "订单准确率"), ("food_safety_score", "食安日检得分"),
    ("critical_temp_fail", "关键温度点异常数"), ("equipment_downtime", "设备停机时长"),
    ("attendance_rate", "出勤达成率"),
    ("late_rate", "迟到率"), ("csat", "顾客满意度"), ("delivery_rating", "外卖平台评分"),
]

# 指标 → 日报里的列名（用于 INDEX/MATCH 取最新日的值）
METRIC_TO_DAILY_COL = {
    "sales_vs_forecast": "销售达成率", "net_sales": "净销售额", "guest_count": "交易笔数",
    "avg_check": "客单价", "items_per_check": "单均件数", "upsell_rate": "附加销售率",
    "delivery_share": "外送占比", "food_cost_pct": "食品成本率", "waste_pct": "废弃率",
    "labor_cost_pct": "人工成本率", "splh": "人时销售SPLH",
    "labor_schedule_variance": "工时排班偏差", "dt_total_time": "得来速总时长",
    "dt_window_time": "得来速取餐窗", "kds_prep_time": "后厨制作时长",
    "front_otd": "前台交付OTD", "delivery_handover_time": "外送出餐时长",
    "speed_hit_rate": "服务速度达标率", "order_accuracy": "订单准确率",
    "complaints_per_1k": "千单投诉数", "food_safety_score": "食安日检得分",
    "critical_temp_fail": "关键温度异常", "holding_time_violation": "保温超时率",
    "cleanliness_score": "清洁巡检得分", "equipment_downtime": "设备停机分钟",
    "attendance_rate": "出勤达成率", "late_rate": "迟到率",
    "station_certification": "岗位认证覆盖率", "crew_turnover_30d": "30日流失率",
    "csat": "顾客满意度", "delivery_rating": "外卖平台评分", "cash_over_short": "现金短溢",
}


def _daily_value_ref(metric_id: str, first: int, last: int, date_cell: str = "$B$3") -> str:
    """从「每日经营日报」按日期取该指标当天的值。"""
    col = DC[METRIC_TO_DAILY_COL[metric_id]]
    return (f"INDEX('{DAILY}'!${col}${first}:${col}${last},"
            f"MATCH({date_cell},'{DAILY}'!$A${first}:$A${last},0))")


def _sheet_dashboard(wb, engine, biz_date, dates) -> None:
    ws = wb.create_sheet("店长看板", 0)
    N = 9
    first, last = 5, 5 + len(dates) - 1
    _title_block(ws, "店长看板 · Store Manager Daily Dashboard",
                 "先看灯，再看今天的三件事，最后才翻矩阵找原因。"
                 "所有数值来自「每日经营日报」，所有红黄绿来自「目标与口径」——两张表都改得动。", N)
    for i, w in enumerate([20, 14, 12, 12, 10, 10, 10, 12, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 顶部：最新营业日 + 一句话结论
    _cell(ws, 3, 1, "最新营业日", bold=True)
    _cell(ws, 3, 2, f"=MAX('{DAILY}'!$A${first}:$A${last})", FMT_DATE, align="center")
    _cell(ws, 3, 4, "当日结论", bold=True)
    ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=N)
    c = _cell(ws, 3, 5,
              '=IF(COUNTIF($D$6:$D$22,"红")>0,"有 "&COUNTIF($D$6:$D$22,"红")&" 项红灯，'
              '按下方「今天的三件事」执行",IF(COUNTIF($D$6:$D$22,"黄")>0,'
              '"无红灯，"&COUNTIF($D$6:$D$22,"黄")&" 项关注","全绿：把时间用在训练和顾客身上"))',
              bold=True)
    c.alignment = Alignment(vertical="center", indent=1)

    # A. 核心 KPI
    r = 5
    _header(ws, r, ["指标", "最新值", "目标", "状态", "绿线", "黄线", "超线带宽", "责任人", "行动卡"])
    r += 1
    kpi_first = r
    for mid, label in KPI_ROWS:
        m = engine.defs[mid]
        fmt = FMT_PCT if m["unit"] == "%" else _fmt_for(m["unit"])
        _cell(ws, r, 1, label, bold=True)
        _cell(ws, r, 2, f"=IFERROR({_daily_value_ref(mid, first, last)},0)", fmt, align="right")
        _cell(ws, r, 3, f"={_lookup('F', mid)}", fmt, align="right")
        _cell(ws, r, 4, _status_formula(f"$B${r}", mid), align="center", bold=True)
        _cell(ws, r, 5, f"={_lookup('G', mid)}", fmt, align="right")
        _cell(ws, r, 6, f"={_lookup('H', mid)}", fmt, align="right")
        _cell(ws, r, 7, _severity_formula(f"$B${r}", mid), FMT_NUM2, align="center")
        _cell(ws, r, 8, f"={_lookup('J', mid)}", align="center")
        _cell(ws, r, 9, f"={_lookup('N', mid)}", align="center")
        r += 1
    kpi_last = r - 1
    _status_cf(ws, f"D{kpi_first}:D{kpi_last}")

    r += 1
    _section(ws, r, "B. 六个家族的灯（绿 / 黄 / 红 计数）", N); r += 1
    _header(ws, r, ["家族", "绿", "黄", "红", "状态", "", "", "", ""]); r += 1
    fam_first = r
    fam_of = {mid: mx.FAMILY_LABEL[engine.defs[mid]["family"]] for mid, _ in KPI_ROWS}
    for fam in ("生意", "速度", "成本", "品质与食安", "人员", "顾客"):
        ids = [i + kpi_first for i, (mid, _) in enumerate(KPI_ROWS) if fam_of[mid] == fam]
        if not ids:
            continue
        rng = ",".join(f"$D${i}" for i in ids)
        _cell(ws, r, 1, fam, bold=True)
        for col, txt in ((2, "绿"), (3, "黄"), (4, "红")):
            _cell(ws, r, col, f'=COUNTIF(({rng}),"{txt}")', FMT_INT, align="center")
        _cell(ws, r, 5, f'=IF(D{r}>0,"红",IF(C{r}>0,"黄","绿"))', align="center", bold=True)
        r += 1
    _status_cf(ws, f"E{fam_first}:E{r-1}")

    # C. 今天的三件事（由优先级矩阵算出，这里引用它，保证只有一份逻辑）
    r += 1
    _section(ws, r, "C. 今天的三件事（取「优先级矩阵」行动分最高的三项，公式动态排名）", N); r += 1
    _header(ws, r, ["#", "指标", "最新值", "状态", "象限", "责任人", "行动卡", "行动分", "第一动作"])
    r += 1
    todo_first = r
    n_pri = len(engine.defs)
    score_rng = f"'优先级矩阵'!$J$5:$J${4 + n_pri}"
    for i in range(1, 4):
        # 第 i 名 = 行动分第 i 大的那一行。用 LARGE+MATCH 做动态排名：
        # 改了阈值/影响度/可控度，这里换的是人，不是数字。
        src_row = f"MATCH(LARGE({score_rng},{i}),{score_rng},0)"
        _cell(ws, r, 1, i, align="center", bold=True)
        for col, src, fmt in ((2, "B", None), (3, "C", None), (4, "E", None), (5, "H", None),
                              (6, "I", None), (7, "K", None), (8, "J", FMT_NUM2), (9, "L", None)):
            _cell(ws, r, col,
                  f"=IFERROR(INDEX('优先级矩阵'!${src}$5:${src}${4 + n_pri},{src_row}),\"\")",
                  fmt, align="center" if col in (4, 5, 6, 7, 8) else None, wrap=(col == 9))
        ws.row_dimensions[r].height = 28
        r += 1
    _status_cf(ws, f"D{todo_first}:D{r-1}")

    r += 1
    _note(ws, r, "「最新值」= 从「每日经营日报」按 B3 的日期取当天的数；"
                 "「目标/绿线/黄线/责任人/行动卡」= 从「目标与口径」按指标ID 回查；"
                 "「超线带宽」= 越过绿线几个黄灯带宽（(绿线−实际)÷|黄线−绿线|），"
                 "用来在不同量纲的指标之间排优先级——不能用相对偏差，"
                 "否则目标 2% 的迟到率会天天霸占第一。", N)
    ws.freeze_panes = "A5"


# ══ Sheet 5 · 优先级矩阵 ══════════════════════════════════════════════════
def _sheet_priority(wb, engine, biz_date, dates) -> None:
    ws = wb.create_sheet("优先级矩阵")
    N = 13
    first, last = 5, 5 + len(dates) - 1
    _title_block(ws, "优先级矩阵 · Priority Matrix（影响度 × 可控度）",
                 "行动分 = 超线带宽 × 影响度 × 可控度 ×（红灯 2 倍）。"
                 "全部是公式：改「目标与口径」的阈值或影响度/可控度，这里的分数和象限会一起变。", N)
    for i, w in enumerate([22, 18, 12, 12, 8, 10, 10, 16, 12, 10, 10, 46, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _header(ws, 4, ["指标ID", "指标", "最新值", "目标", "状态", "影响度", "可控度",
                    "象限", "责任人", "行动分", "行动卡", "第一动作", "排名"])

    # 排序在 Python 里做（Excel 的 SORT 不可用），但分数本身仍是公式
    pri = mx.priority(engine, biz_date, include_green=True)
    items = sorted(pri["items"], key=lambda x: -x["score"])
    order_ids = [it["id"] for it in items]

    playbook_action = {
        "PB-01": "分清是客流少还是客单少：GC 跌查外部与曝光，客单跌走附加销售",
        "PB-02": "拉附加销售率 × 收银员，训练员跟班 2 小时；只主推 1 个品",
        "PB-03": "全时段慢=打包流程问题（上报）；只高峰慢=给外送单独留打包位",
        "PB-04": "先分段定位（点单/排队/取餐窗/制作），再看是结构问题还是产能问题",
        "PB-05": "按差错类型归因；同一 tag 一周 3 次以上 = 流程缺陷，进周复盘",
        "PB-06": "分成品/原料/保温三类：成品高改批量，原料高查先进先出",
        "PB-07": "看 SPLH 与速度的组合：都差 = 排班结构错，调岗位不是加人头",
        "PB-08": "定位班次与收银员，复核交接记录，当天闭环不隔夜",
        "PB-09": "关键项 fail 立即停止该产品出品，整改复检通过才恢复",
        "PB-10": "分区看：洗手间失分优先（顾客对干净的判断在那里）",
        "PB-11": "先决定是否停售并同步到所有点单渠道（含外送平台），再报修",
        "PB-12": "按司龄段/岗位下钻；<90 天流失高 = 新人被丢在高峰上",
    }

    r = 5
    for mid in order_ids:
        m = engine.defs[mid]
        fmt = FMT_PCT if m["unit"] == "%" else _fmt_for(m["unit"])
        band = (r % 2 == 0)
        _cell(ws, r, 1, mid, band=band)
        _cell(ws, r, 2, m["name_cn"], bold=True, band=band)
        _cell(ws, r, 3, f"=IFERROR({_daily_value_ref(mid, first, last, date_cell='店长看板!$B$3')},0)",
              fmt, align="right")
        _cell(ws, r, 4, f"={_lookup('F', mid)}", fmt, align="right")
        _cell(ws, r, 5, _status_formula(f"$C${r}", mid), align="center", bold=True)
        _cell(ws, r, 6, f"={_lookup('L', mid)}", FMT_NUM2, align="center", band=band)
        _cell(ws, r, 7, f"={_lookup('M', mid)}", FMT_NUM2, align="center", band=band)
        _cell(ws, r, 8, f'=IF($F${r}>=0.7,IF($G${r}>=0.7,"Q1 今天就做","Q2 上报求援"),'
                        f'IF($G${r}>=0.7,"Q3 顺手做掉","Q4 记录观察"))', align="center", band=band)
        _cell(ws, r, 9, f"={_lookup('J', mid)}", align="center", band=band)
        _cell(ws, r, 10, f'=IFERROR(ROUND({_severity_formula(f"$C${r}", mid)[1:]}'
                         f'*$F${r}*$G${r}*IF($E${r}="红",2,1),2),0)', FMT_NUM2, align="center", bold=True)
        _cell(ws, r, 11, f"={_lookup('N', mid)}", align="center", band=band)
        _cell(ws, r, 12, playbook_action.get(m["playbook"], ""), wrap=True, band=band)
        _cell(ws, r, 13, f"=RANK($J${r},$J$5:$J${5 + len(order_ids) - 1},0)",
              FMT_INT, align="center", band=band)
        ws.row_dimensions[r].height = 26
        r += 1
    lastr = r - 1
    _status_cf(ws, f"E5:E{lastr}")
    for txt, bg in (("Q1 今天就做", "FFE0E0"), ("Q2 上报求援", "FFF0D8"),
                    ("Q3 顺手做掉", "E3F0FB"), ("Q4 记录观察", "F0F0F0")):
        ws.conditional_formatting.add(f"H5:H{lastr}", CellIsRule(
            operator="equal", formula=[f'"{txt}"'], fill=PatternFill("solid", bgColor=bg)))
    ws.conditional_formatting.add(f"J5:J{lastr}", ColorScaleRule(
        start_type="min", start_color="FFFFFF", end_type="max", end_color="F5A3A3"))
    ws.conditional_formatting.add(f"M5:M{lastr}", CellIsRule(
        operator="lessThanOrEqual", formula=["3"],
        fill=PatternFill("solid", bgColor=RED), font=Font(name=FONT, size=9, bold=True, color=RED_INK)))
    ws.freeze_panes = "C5"

    r = lastr + 2
    _section(ws, r, "四象限的含义", N); r += 1
    _header(ws, r, ["象限", "影响度", "可控度", "含义", "", "", "", "", "", "", ""]); r += 1
    for q, imp, ctl, mean in (
        ("Q1 今天就做", "高", "高", "这就是今天班前会的内容"),
        ("Q2 上报求援", "高", "低", "影响大但不在你手上——找营运经理/供应商/总部，别自己耗一整天"),
        ("Q3 顺手做掉", "低", "高", "一动就好，交给值班经理清单化处理"),
        ("Q4 记录观察", "低", "低", "别在今天的会上讲，记进周复盘看趋势"),
    ):
        _cell(ws, r, 1, q, bold=True)
        _cell(ws, r, 2, imp, align="center")
        _cell(ws, r, 3, ctl, align="center")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=N)
        _cell(ws, r, 4, mean, wrap=True)
        r += 1
    r += 1
    _note(ws, r, "行序在生成时按行动分排好，但「排名」列（M）是公式：改了阈值或影响度/可控度，"
                 "排名会立刻重算，「店长看板」的三件事也跟着换人——它取的是排名前三，不是前三行。"
                 "想让行序也跟着动，用「数据 → 排序」按 J 列降序即可。", N)


# ══ Sheet 6 · 时段矩阵 ════════════════════════════════════════════════════
INTERVAL_METRICS = ["sales_vs_forecast", "speed_hit_rate", "dt_total_time", "kds_prep_time", "splh"]


def _sheet_interval(wb, engine, biz_date) -> None:
    ws = wb.create_sheet("时段矩阵")
    g = mx.interval_grid(engine, biz_date, INTERVAL_METRICS)
    ivs = g["intervals"]
    N = len(ivs) + 2
    _title_block(ws, "时段矩阵 · Interval Matrix（15 分钟 × 指标）",
                 f"{biz_date} 全天，每格 15 分钟。看法只有一条：单点飘红是噪声，"
                 "连成一片才是问题——下方自动列出了连续 2 格以上的告警窗口。", min(N, 40))
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 9
    for i in range(3, N + 1):
        ws.column_dimensions[get_column_letter(i)].width = 5.4

    hdr = ["指标", "单位"] + ivs
    _header(ws, 4, hdr)
    for i, iv in enumerate(ivs, start=3):
        ws.cell(4, i).alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=90)
    ws.row_dimensions[4].height = 46

    r = 5
    for mid in INTERVAL_METRICS:
        m = engine.defs[mid]
        _cell(ws, r, 1, m["name_cn"], bold=True)
        _cell(ws, r, 2, m["unit"], align="center")
        for i, cell in enumerate(g["grid"][mid], start=3):
            v = _to_sheet(cell["value"], m["unit"])
            c = _cell(ws, r, i, round(v, 4) if v is not None else None,
                      FMT_PCT if m["unit"] == "%" else FMT_INT, align="center", size=7.5)
            bg = {"green": None, "amber": AMBER, "red": RED, "na": None}[cell["status"]]
            if bg:
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name=FONT, size=7.5, bold=(cell["status"] == "red"),
                              color=RED_INK if cell["status"] == "red" else AMBER_INK)
        ws.row_dimensions[r].height = 20
        r += 1

    r += 1
    _section(ws, r, "连续告警窗口（连续 ≥2 个 15 分钟非绿）", min(N, 40)); r += 1
    _header(ws, r, ["指标", "从", "到", "连续格数", "均值", "最差", "含红灯", "→ 说明"]
            + [""] * max(0, min(N, 40) - 8))
    for i, w in enumerate([18, 9, 9, 10, 10, 10, 9, 52], start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(
            ws.column_dimensions[get_column_letter(i)].width or 0, w) if i <= 2 else w
    r += 1
    any_window = False
    for mid in INTERVAL_METRICS:
        m = engine.defs[mid]
        fmt = FMT_PCT if m["unit"] == "%" else FMT_INT
        for w in mx.problem_windows(engine, biz_date, mid):
            any_window = True
            _cell(ws, r, 1, m["name_cn"], bold=True)
            _cell(ws, r, 2, w["from"], align="center")
            _cell(ws, r, 3, w["to"], align="center")
            _cell(ws, r, 4, w["slots"], FMT_INT, align="center")
            _cell(ws, r, 5, _to_sheet(w["avg"], m["unit"]), fmt, align="right")
            _cell(ws, r, 6, _to_sheet(w["worst"], m["unit"]), fmt, align="right")
            _cell(ws, r, 7, "是" if w["has_red"] else "否", align="center")
            _cell(ws, r, 8, "连片告警，去「交叉矩阵」确认是某个渠道的结构问题还是整段的产能问题",
                  wrap=True)
            r += 1
    if not any_window:
        _cell(ws, r, 1, "今天没有连续告警窗口", bold=True)
        r += 1
    r += 1
    _note(ws, r, "为什么用 15 分钟而不是小时：排班和备货的最小调整单位就是 15 分钟。"
                 "按小时看，一个 20 分钟的塌陷会被两边的好数据稀释掉，什么也看不见。", min(N, 40))
    ws.freeze_panes = "C5"


# ══ Sheet 7 · 交叉矩阵 ════════════════════════════════════════════════════
def _sheet_cross(wb, engine, biz_date) -> None:
    ws = wb.create_sheet("交叉矩阵")
    p = engine.profile
    ch_cn = {c["id"]: c["name_cn"] for c in p["channels"]}
    dp_cn = {d["id"]: d["name_cn"] for d in p["dayparts"]}
    N = 8
    _title_block(ws, "交叉矩阵 · Cross Matrix（时段 × 渠道）",
                 "一张表同时读出两种问题：某一列全红 = 结构问题（那个渠道的流程本身有毛病，"
                 "加人没用）；某一行全红 = 产能问题（那个时段人不够，排班能解决）。"
                 "两种问题的动作、责任人、时间尺度都不一样。", N)
    for i, w in enumerate([14, 13, 13, 13, 13, 13, 12, 40], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 4
    for mid in ("speed_hit_rate", "net_sales", "waste_pct"):
        m = engine.defs[mid]
        c = mx.cross(engine, biz_date, mid, "daypart", "channel")
        if c["note"]:
            continue
        fmt = FMT_PCT if m["unit"] == "%" else _fmt_for(m["unit"])
        _section(ws, r, f"{m['name_cn']}（{m['unit']}） · 时段 × 渠道", N); r += 1
        _header(ws, r, ["时段"] + [ch_cn.get(b, b) for b in c["b_values"]] + ["行合计/均值", "读法"])
        r += 1
        top = r
        for row in c["cells"]:
            _cell(ws, r, 1, dp_cn.get(row[0]["a"], row[0]["a"]), bold=True)
            for i, cell in enumerate(row, start=2):
                v = _to_sheet(cell["value"], m["unit"])
                cc = _cell(ws, r, i, round(v, 4) if v is not None else None, fmt, align="right")
                bg = {"green": GREEN, "amber": AMBER, "red": RED, "na": None}[cell["status"]]
                if bg:
                    cc.fill = PatternFill("solid", fgColor=bg)
                    cc.font = Font(name=FONT, size=9, bold=(cell["status"] == "red"),
                                   color={"green": GREEN_INK, "amber": AMBER_INK,
                                          "red": RED_INK}[cell["status"]])
            ncol = len(c["b_values"])
            agg = "SUM" if m.get("aggregation") == "sum" else "AVERAGE"
            _cell(ws, r, ncol + 2, f"=IFERROR({agg}(B{r}:{get_column_letter(ncol+1)}{r}),0)",
                  fmt, align="right", bold=True)
            r += 1
        ncol = len(c["b_values"])
        _cell(ws, r, 1, "列合计/均值", bold=True)
        for i in range(2, ncol + 2):
            col = get_column_letter(i)
            agg = "SUM" if m.get("aggregation") == "sum" else "AVERAGE"
            _cell(ws, r, i, f"=IFERROR({agg}({col}{top}:{col}{r-1}),0)", fmt,
                  align="right", bold=True)
        r += 1
        if c["worst"]:
            w = c["worst"][0]
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
            _cell(ws, r, 1, f"→ 最拖后腿的格子：{dp_cn.get(w['a'], w['a'])} × "
                            f"{ch_cn.get(w['b'], w['b'])}（拖累度 {w['contribution']:.1f}）",
                  bold=True)
            r += 1
        r += 1

    _note(ws, r, "红黄绿的判定：率型指标直接对「目标与口径」的绿黄线；"
                 "量型指标（净销售）改判「对该切片预测的达成率」——"
                 "拿全店目标 62,000 去判晚市的 4,761 元，整张表会永远是红的，那就没有信息了。"
                 "详见 docs/03 第 3.3 节。", N)
    ws.freeze_panes = "B5"


# ══ Sheet 8 · 归因矩阵 ════════════════════════════════════════════════════
def _sheet_attribution(wb, engine, biz_date) -> None:
    ws = wb.create_sheet("归因矩阵")
    N = 9
    _title_block(ws, "归因矩阵 · Attribution Matrix（指标 × 维度）",
                 "排序列是「拖累度」，不是指标值本身。外送废弃率再高，只占 3% 的量也不值得先动；"
                 "堂食占 70%，压 0.2 个百分点就顶外送归零的两倍。", N)
    for i, w in enumerate([20, 16, 12, 10, 12, 12, 11, 12, 34], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    pri = mx.priority(engine, biz_date)
    focus = [it["id"] for it in pri["items"][:6]] or ["speed_hit_rate", "waste_pct"]

    r = 4
    for mid in focus:
        m = engine.defs[mid]
        fmt = FMT_PCT if m["unit"] == "%" else _fmt_for(m["unit"])
        for dim in m.get("dims", [])[:2]:
            a = mx.attribution(engine, biz_date, mid, dim)
            rows = [x for x in a["rows"] if x["value"] is not None]
            if not rows:
                continue
            _section(ws, r, f"{m['name_cn']} × {a['dim_cn']}　"
                            f"（全店 {rp.fmt(a['store_value'], m['unit'])}，"
                            f"贡献度口径：{a['contribution_basis']}）", N)
            r += 1
            _header(ws, r, [a["dim_cn"], "值", "状态", "量占比", "该格预算", a["gap_label"],
                            "拖累度", "累计占比", "读法"])
            r += 1
            top = r
            total = sum(max(0.0, x["contribution"] or 0) for x in rows) or 1.0
            running = 0.0
            for x in rows:
                running += max(0.0, x["contribution"] or 0)
                _cell(ws, r, 1, str(x["dim_value"]), bold=True)
                v = _to_sheet(x["value"], m["unit"])
                cc = _cell(ws, r, 2, round(v, 4) if v is not None else None, fmt, align="right")
                st = _cell(ws, r, 3, {"green": "绿", "amber": "黄", "red": "红", "na": "—"}[x["status"]],
                           align="center", bold=True)
                _cell(ws, r, 4, x["weight"], FMT_PCT, align="right")
                # 该格应得的预算：全店目标 × 量占比（只对全店分母口径的指标成立）
                if m.get("denominator") == "store_day":
                    _cell(ws, r, 5, f"={_lookup('F', mid)}*$D${r}", fmt, align="right")
                else:
                    _cell(ws, r, 5, f"={_lookup('F', mid)}", fmt, align="right")
                _cell(ws, r, 6, _to_sheet(x["gap"], m["unit"]) if x["gap"] is not None else None,
                      fmt, align="right")
                _cell(ws, r, 7, round(x["contribution"] or 0, 2), FMT_NUM2, align="right", bold=True)
                _cell(ws, r, 8, running / total, FMT_PCT, align="right")
                _cell(ws, r, 9, "", wrap=True)
                r += 1
            _status_cf(ws, f"C{top}:C{r-1}")
            ws.conditional_formatting.add(f"G{top}:G{r-1}", ColorScaleRule(
                start_type="min", start_color="FFFFFF", end_type="max", end_color="F5A3A3"))
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
            head = rows[0]
            _cell(ws, r, 1, f"→ 先动 {head['dim_value']}：它占了拖累度的 "
                            f"{max(0.0, head['contribution'] or 0)/total*100:.0f}%"
                            f"（量占比 {head['weight']*100:.0f}%）", bold=True)
            r += 2

    _note(ws, r, "「该格预算」= 全店目标 × 该格量占比，只对全店分母口径的指标（废弃率、食品成本率、"
                 "千单投诉数）成立：全店废弃率目标 0.8%、午市占 34% 的销售，午市的预算就是 0.27 个百分点。"
                 "各格预算加起来正好等于全店目标，所以这样判色是自洽的。"
                 "切片分母口径的指标（速度、人工率、客单价）目标与切片大小无关，直接用全店目标。", N)


# ══ Sheet 9 · 报表清单 ════════════════════════════════════════════════════
def _sheet_reports(wb, engine) -> None:
    ws = wb.create_sheet("报表清单")
    defs = engine.defs
    N = 9
    _title_block(ws, "报表清单 · Manager Report Package",
                 "一张报表如果回答不了「看完要做什么」，它就不该出现在这个清单里。"
                 "所以最后一栏是动作，不是描述。全天合计不超过 75 分钟——"
                 "超过这个数，报表就开始和营运抢时间了。", N)
    for i, w in enumerate([8, 24, 18, 8, 10, 34, 32, 30, 38], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _section(ws, 4, "A. 每日报表（按一天的时间轴）", N)
    _header(ws, 5, ["编号", "报表", "什么时候", "分钟", "谁看", "回答什么问题",
                    "关键指标", "数据来源表", "看完做什么"])
    r = 6
    first = r
    for rep in rp.DAILY_REPORTS:
        band = (r % 2 == 0)
        _cell(ws, r, 1, rep["code"], align="center", bold=True, band=band)
        _cell(ws, r, 2, rep["name"], bold=True, wrap=True, band=band)
        _cell(ws, r, 3, rep["when"], wrap=True, band=band)
        _cell(ws, r, 4, rep["minutes"], FMT_INT, align="center", band=band)
        _cell(ws, r, 5, mx.OWNER_LABEL[rep["owner"]], align="center", band=band)
        _cell(ws, r, 6, rep["question"], wrap=True, band=band)
        _cell(ws, r, 7, "、".join(defs[m]["name_cn"] for m in rep["metrics"]), wrap=True, band=band)
        _cell(ws, r, 8, "、".join(rep["tables"]), wrap=True, band=band)
        _cell(ws, r, 9, rep["action"], wrap=True, band=band)
        ws.row_dimensions[r].height = 42
        r += 1
    _cell(ws, r, 2, "合计", bold=True)
    _cell(ws, r, 4, f"=SUM(D{first}:D{r-1})", FMT_INT, align="center", bold=True)
    _cell(ws, r, 6, "全天报表时间上限：75 分钟", bold=True)
    r += 2

    _section(ws, r, "B. 每周报表", N); r += 1
    _header(ws, r, ["编号", "报表", "频率", "", "谁看", "回答什么问题", "关键指标", "", ""])
    r += 1
    for rep in rp.WEEKLY_REPORTS:
        _cell(ws, r, 1, rep["code"], align="center", bold=True)
        _cell(ws, r, 2, rep["name"], bold=True, wrap=True)
        _cell(ws, r, 3, "每周", align="center")
        _cell(ws, r, 5, mx.OWNER_LABEL[rep["owner"]], align="center")
        _cell(ws, r, 6, rep["question"], wrap=True)
        _cell(ws, r, 7, "、".join(defs[m]["name_cn"] for m in rep["metrics"]), wrap=True)
        ws.row_dimensions[r].height = 32
        r += 1

    r += 1
    _section(ws, r, "C. 责任矩阵：谁 × 什么节奏 × 盯什么", N); r += 1
    own = mx.ownership(engine)
    _header(ws, r, ["责任人"] + [mx.CADENCE_LABEL[c] for c in own["cadences"]] + [""] * 5)
    r += 1
    for o in own["owners"]:
        if o not in own["grid"]:
            continue
        _cell(ws, r, 1, mx.OWNER_LABEL[o], bold=True)
        for i, cad in enumerate(own["cadences"], start=2):
            items = own["grid"][o].get(cad, [])
            _cell(ws, r, i, "、".join(items) if items else "—", wrap=True)
        ws.row_dimensions[r].height = 40
        r += 1
    r += 1
    _note(ws, r, "责任矩阵一年也不改几次，但它是「这本工作簿到底归谁管」的唯一答案，"
                 "也是新店长交接时第一份要看的东西。", N)
    ws.freeze_panes = "C6"


# ══ Sheet 10 · 数据字典 ═══════════════════════════════════════════════════
def _sheet_dictionary(wb) -> None:
    ws = wb.create_sheet("数据字典")
    N = 9
    _title_block(ws, "数据字典 · Source Tables",
                 "一家门店每天真正在用的 30 张表。现实是它们分散在 6~8 个互不相通的系统里；"
                 "把它们对齐到同一个 store_id + 营业日，是做任何报表的前提，"
                 "也是最容易出错的地方（跨零点的夜宵单算哪一天）。", N)
    for i, w in enumerate([24, 18, 10, 8, 20, 26, 16, 14, 46], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _header(ws, 4, ["表名", "中文名", "业务域", "层级", "来源系统", "粒度", "刷新频率",
                    "日增行数", "为什么需要它"])
    r = 5
    for domain, tables in schema.by_domain().items():
        for t in tables:
            band = (r % 2 == 0)
            _cell(ws, r, 1, t["name"], bold=True, band=band)
            _cell(ws, r, 2, t["cn"], band=band)
            _cell(ws, r, 3, domain, align="center", band=band)
            _cell(ws, r, 4, t["layer"], align="center", band=band)
            _cell(ws, r, 5, t["owner_system"], wrap=True, band=band)
            _cell(ws, r, 6, t["grain"], wrap=True, band=band)
            _cell(ws, r, 7, t["refresh"], wrap=True, band=band)
            _cell(ws, r, 8, t["rows_per_day"], align="center", band=band)
            _cell(ws, r, 9, t["why"], wrap=True, band=band)
            ws.row_dimensions[r].height = 30
            r += 1
    ws.freeze_panes = "C5"
    _add_table(ws, "SourceTables", f"A4:I{r-1}")


# ══ Sheet 11 · 行动卡 ═════════════════════════════════════════════════════
PLAYBOOKS = [
    ("PB-01", "销售 / 客流不达标", "先分清是「人少」还是「人均少」：看交易笔数和客单价哪个先塌",
     "客流少 → 查天气/施工/竞品/平台曝光/点餐机故障；客单少 → 走 PB-02；两个都跌 → 单时段跌是具体事件，全时段跌是商圈或平台层面",
     "次日同时段同渠道回到预测的 ±3% 以内"),
    ("PB-02", "客单价 / 附加销售不足", "人工点单低、点餐机正常 = 人的问题；两个都低 = 产品结构问题（上报）",
     "拉附加销售率 × 收银员找出后 30%；班前会只指定 1 个主推品（推三个等于不推）；训练员跟班 2 小时",
     "三天内后 30% 的人追上门店均值的 80%"),
    ("PB-03", "外送时效 / 评分下滑", "用堂食的办法救外送通常无效——它是两套流程",
     "全时段慢 = 打包动线与堂食抢人（跨部门，上报）；只高峰慢 = 高峰给外送单独留人；某平台慢 = 联系平台商务；差评集中漏餐 = 双人核对 + 封签",
     "出餐时长回到目标线连续 3 天；评分是滞后指标，一到两周后才跟上"),
    ("PB-04", "服务速度不达标", "最需要先诊断再动手的一张卡。直接「加人」是最常见也最贵的错误答案",
     "① 分段定位：点单段长→加点单员/开第二车道；取餐窗长→后厨备餐没跟上，加人到理货（不是点单）；制作段长→查设备/批量/新手 ② 分清结构问题（某渠道全时段红）还是产能问题（某时段所有渠道红）",
     "目标时段达标率回绿，且 SPLH 没有明显下降（否则是靠堆人换的，不可持续）"),
    ("PB-05", "订单差错 / 投诉上升", "先看是不是同一件事在重复发生",
     "集中在一种差错类型 → 改流程；集中在一个时段 → 走 PB-04；集中在一个渠道 → 查该渠道的交接环节；分散无规律 → 看岗位认证覆盖率",
     "千单投诉数连续 5 天在绿线内，且同类差错不再出现在 top 3"),
    ("PB-06", "废弃 / 食品成本超标", "先分清废弃的三种类型，动作完全不同",
     "成品废弃（做多了）→ 减小批量、缩短备货提前量；原料废弃（过期掉落）→ 查先进先出、调订货量；保温超时 → 高峰前按预测分批做。食品成本高但废弃正常 → 问题在盘点差异：查配方/打料/收货/废弃漏登记",
     "次日同时段废弃回绿；一周后食品成本率跟上"),
    ("PB-07", "人工成本 / 排班效率", "看 SPLH 与速度的组合，这是判断根因的关键",
     "SPLH 低 + 速度差 = 排班结构错（人站错岗），调岗位不是加人头；SPLH 低 + 速度好 = 人排多了，削平峰保高峰；SPLH 高 + 速度差 = 真的需要加人；两个都好 = 记下这天的排班曲线做模板",
     "目标时段 SPLH 回绿，且速度达标率不下降"),
    ("PB-08", "现金短溢", "金额小但性质是内控问题，不按金额大小排优先级",
     "定位到班次和收银员 → 复核交接记录与监控 → 当天闭环不隔夜 → 同一人连续出现转人事流程",
     "当日闭环并留下书面记录"),
    ("PB-09", "食安失分 / 温度异常", "这张卡没有「先记着」这个选项",
     "关键项 fail = 立即停止该产品出品，整改复检通过才恢复；温度超标 = 判断受影响批次范围并按标准处置 → 报修 → 加密监测；冷藏/冷冻库超标可能触发全库处置",
     "复检通过 + 连续 3 天该项无 fail；同一项一周 fail 两次必须上报"),
    ("PB-10", "清洁 / 巡店失分", "顾客对「干净」的判断在洗手间和餐桌，不在厨房",
     "洗手间失分 → 高峰期每小时巡检一次（投入产出比最高）；大堂失分 → 高峰收餐人力不足，走 PB-07；外围/车道失分 → 通常是排班里根本没有这个岗位",
     "分区得分全部回到绿线，且洗手间不再是最低分区"),
    ("PB-11", "设备停机", "先决定停售范围，再谈修",
     "关键设备停机 → 立刻决定哪些产品下架并同步到所有点单渠道（最容易忘的是外送平台：平台还在卖但门店做不出来，是差评的最大来源之一）→ 报修 → 同一设备一月两次故障走更换评估",
     "恢复出品 + 工单闭环；周会看一次保养到期清单"),
    ("PB-12", "人员流失 / 认证不足", "唯一一张不看日数据的卡——日维度上这两个指标全是噪声",
     "流失集中在 <90 天 → 入职培训和排班体验的问题（新人被丢在高峰上是最常见原因）；流失分布在 >1 年 → 发展通道或薪酬，超出门店范围，上报；某岗位认证覆盖率长期低 → 训练排期没覆盖到它，而它往往正是速度和废弃问题的源头",
     "进周会第一页跟踪；这两个指标是速度/废弃/食安/差错红灯的共同上游"),
]


def _sheet_playbooks(wb) -> None:
    ws = wb.create_sheet("行动卡")
    N = 5
    _title_block(ws, "行动卡 · Action Playbooks",
                 "红灯亮了做什么。每张卡的结构固定：先判断根因（同一个红灯，根因不同则动作相反，"
                 "猜错了比不做还糟），再给动作，最后给验证方式。", N)
    for i, w in enumerate([9, 22, 40, 62, 34], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _header(ws, 4, ["编号", "适用情况", "先判断什么（诊断）", "动作", "怎么验证"])
    r = 5
    for code, title, diag, action, verify in PLAYBOOKS:
        band = (r % 2 == 0)
        _cell(ws, r, 1, code, align="center", bold=True, band=band)
        _cell(ws, r, 2, title, bold=True, wrap=True, band=band)
        _cell(ws, r, 3, diag, wrap=True, band=band)
        _cell(ws, r, 4, action, wrap=True, band=band)
        _cell(ws, r, 5, verify, wrap=True, band=band)
        ws.row_dimensions[r].height = 62
        r += 1
    ws.freeze_panes = "B5"
    _note(ws, r + 1, "行动卡编号挂在「目标与口径」的最后一列，所以「店长看板」和「优先级矩阵」"
                     "上的每个红灯都会直接告诉你翻哪一张卡。", N)


# ══ 组装 ═════════════════════════════════════════════════════════════════
def build_workbook(engine: MetricEngine, biz_date: date, days: int, out_path) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    # 顺序：目标与口径必须先建（它决定 T_FIRST/T_LAST，后面的公式要用）
    _sheet_targets(wb, engine)
    dates = _sheet_daily(wb, engine, biz_date, days)
    _sheet_dashboard(wb, engine, biz_date, dates)      # 插到最前
    _sheet_priority(wb, engine, biz_date, dates)
    _sheet_interval(wb, engine, biz_date)
    _sheet_cross(wb, engine, biz_date)
    _sheet_attribution(wb, engine, biz_date)
    _sheet_reports(wb, engine)
    _sheet_playbooks(wb)
    _sheet_dictionary(wb)
    _sheet_readme(wb, engine, biz_date, days)

    # 排成阅读顺序
    desired = ["店长看板", "每日经营日报", "优先级矩阵", "时段矩阵", "交叉矩阵",
               "归因矩阵", "报表清单", "行动卡", TARGETS, "数据字典", "使用说明"]
    wb._sheets = [wb[name] for name in desired if name in wb.sheetnames]

    tab_colors = {"店长看板": BRAND, "每日经营日报": NAVY, "优先级矩阵": BRAND,
                  "时段矩阵": "4A6572", "交叉矩阵": "4A6572", "归因矩阵": "4A6572",
                  "报表清单": "6B7280", "行动卡": "6B7280", TARGETS: "8A6100",
                  "数据字典": "9AA3AE", "使用说明": "9AA3AE"}
    for name, color in tab_colors.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color

    wb.properties.title = "门店日运营报表包"
    wb.properties.creator = "mcd_store_ops"
    wb.save(out_path)
    return str(out_path)
