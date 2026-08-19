"""验证 Excel 工作簿里的公式算出来的数，等于指标引擎算出来的数。

recalc 报 0 错误只证明公式**能求值**，不证明它们**是对的**——
一个差一行的 INDEX 范围会得到一份干净无错、但数字全错的工作簿。
所以必须逐项比对。

前置：
    python -m mcdops.cli excel
    python <xlsx-skill>/scripts/recalc.py output/store_ops_package.xlsx 360

运行：
    PYTHONPATH=src python3 tests/verify_excel.py
退出码 0 = 全部一致。
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from mcdops import matrix as mx                                  # noqa: E402
from mcdops.excel import KPI_ROWS                                # noqa: E402
from mcdops.metrics import MetricEngine, load_metrics, load_profile  # noqa: E402
from mcdops.simulate import simulate_period                      # noqa: E402

XLSX = Path(__file__).resolve().parents[1] / "output" / "store_ops_package.xlsx"
BIZ_DATE = date(2026, 8, 18)
DAYS = 14

STATUS_CN = {"green": "绿", "amber": "黄", "red": "红", "na": "—"}
# 工作簿里数值按显示需要做了 round（秒 2 位、评分 4 位），比这个宽松即可
TOL_REL = 1e-4
TOL_ABS = 1e-6

failures: list[str] = []
checks = 0


def close(a, b) -> bool:
    if a is None or b is None:
        return a is b
    return abs(a - b) <= max(TOL_ABS, abs(b) * TOL_REL)


def check(ok: bool, label: str, got=None, want=None) -> None:
    global checks
    checks += 1
    if not ok:
        failures.append(f"{label}: Excel={got!r} 引擎={want!r}")


def main() -> int:
    if not XLSX.exists():
        print(f"✗ 找不到 {XLSX}，先跑 python -m mcdops.cli excel")
        return 1

    profile = load_profile()
    engine = MetricEngine(simulate_period(profile, BIZ_DATE, days=DAYS), profile, load_metrics())
    truth = {r["id"]: r for r in engine.scorecard(BIZ_DATE)}

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    if wb["店长看板"]["B3"].value is None:
        print("✗ 公式没有缓存值——先跑 recalc.py")
        return 1

    # ── 1. 店长看板：最新值 + 状态
    ws = wb["店长看板"]
    got_date = ws["B3"].value
    check(got_date.date() == BIZ_DATE if hasattr(got_date, "date") else got_date == BIZ_DATE,
          "看板 B3 最新营业日", got_date, BIZ_DATE)
    for i, (mid, label) in enumerate(KPI_ROWS):
        r = 6 + i
        m = engine.defs[mid]
        want = truth[mid]["value"]
        if m["unit"] == "%" and want is not None:
            want /= 100.0
        check(close(ws.cell(r, 2).value, want), f"看板·{label}·值", ws.cell(r, 2).value, want)
        check(ws.cell(r, 4).value == STATUS_CN[truth[mid]["status"]],
              f"看板·{label}·状态", ws.cell(r, 4).value, STATUS_CN[truth[mid]["status"]])

    # ── 2. 阈值三方一致：看板 ← 目标与口径 ← metrics.json
    tg = wb["目标与口径"]
    trow = {tg.cell(r, 1).value: r for r in range(5, tg.max_row + 1) if tg.cell(r, 1).value}
    for i, (mid, label) in enumerate(KPI_ROWS):
        r, tr = 6 + i, trow[mid]
        m = engine.defs[mid]
        f = 100.0 if m["unit"] == "%" else 1.0
        for dash_col, tgt_col, key in ((3, 6, "target"), (5, 7, "green"), (6, 8, "amber")):
            a, b = ws.cell(r, dash_col).value, tg.cell(tr, tgt_col).value
            check(close(a, b), f"阈值回查·{label}·{key}（看板↔目标与口径）", a, b)
            check(close(b, m[key] / f), f"阈值源头·{label}·{key}（目标与口径↔metrics.json）",
                  b, m[key] / f)

    # ── 3. 日报的公式列 = 引擎指标
    ws = wb["每日经营日报"]
    hdr = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    row = next(r for r in range(5, ws.max_row + 1)
               if ws.cell(r, 1).value and ws.cell(r, 1).value.date() == BIZ_DATE)
    for col, mid in [
        ("销售达成率", "sales_vs_forecast"), ("客单价", "avg_check"),
        ("单均件数", "items_per_check"), ("附加销售率", "upsell_rate"),
        ("外送占比", "delivery_share"), ("食品成本率", "food_cost_pct"),
        ("废弃率", "waste_pct"), ("人工成本率", "labor_cost_pct"),
        ("人时销售SPLH", "splh"), ("工时排班偏差", "labor_schedule_variance"),
        ("服务速度达标率", "speed_hit_rate"), ("订单准确率", "order_accuracy"),
        ("千单投诉数", "complaints_per_1k"), ("保温超时率", "holding_time_violation"),
        ("出勤达成率", "attendance_rate"), ("迟到率", "late_rate"),
        ("岗位认证覆盖率", "station_certification"), ("顾客满意度", "csat"),
    ]:
        want = engine.compute(BIZ_DATE, [mid])[mid]
        if engine.defs[mid]["unit"] == "%" and want is not None:
            want /= 100.0
        check(close(ws.cell(row, hdr[col]).value, want),
              f"日报公式列·{col}", ws.cell(row, hdr[col]).value, want)

    # ── 4. 优先级矩阵：行序、行动分、象限
    ws = wb["优先级矩阵"]
    items = sorted(mx.priority(engine, BIZ_DATE, include_green=True)["items"],
                   key=lambda x: -x["score"])
    for i, it in enumerate(items[:12]):
        r = 5 + i
        check(ws.cell(r, 1).value == it["id"], f"优先级·第{i+1}行·指标ID",
              ws.cell(r, 1).value, it["id"])
        check(abs((ws.cell(r, 10).value or 0) - it["score"]) <= 0.011,
              f"优先级·{it['id']}·行动分", ws.cell(r, 10).value, it["score"])
        check(ws.cell(r, 8).value == it["quadrant"], f"优先级·{it['id']}·象限",
              ws.cell(r, 8).value, it["quadrant"])

    # ── 5. 时段矩阵：抽查每个指标的若干格
    ws = wb["时段矩阵"]
    ivs = [ws.cell(4, c).value for c in range(3, ws.max_column + 1) if ws.cell(4, c).value]
    from mcdops.excel import INTERVAL_METRICS
    for mi, mid in enumerate(INTERVAL_METRICS):
        m = engine.defs[mid]
        for iv in (ivs[0], ivs[len(ivs) // 2], ivs[-1]):
            want = engine.compute(BIZ_DATE, [mid], "interval15", iv)[mid]
            if want is None:
                continue
            if m["unit"] == "%":
                want /= 100.0
            got = ws.cell(5 + mi, 3 + ivs.index(iv)).value
            check(close(got, want), f"时段矩阵·{m['name_cn']}·{iv}", got, want)

    # ── 6. 交叉矩阵：速度达标率的每一格
    ws = wb["交叉矩阵"]
    c = mx.cross(engine, BIZ_DATE, "speed_hit_rate", "daypart", "channel")
    top = next(r for r in range(4, 40) if ws.cell(r, 1).value == "早餐")
    for i, crow in enumerate(c["cells"]):
        for j, cell in enumerate(crow):
            if cell["value"] is None:
                continue
            got = ws.cell(top + i, 2 + j).value
            check(close(got, cell["value"] / 100.0),
                  f"交叉矩阵·{cell['a']}×{cell['b']}", got, cell["value"] / 100.0)

    # ── 报告
    print(f"\n核对 {checks} 项 · 通过 {checks - len(failures)} · 失败 {len(failures)}\n")
    for f in failures:
        print(f"  ✗ {f}")
    if not failures:
        print("  ✓ Excel 工作簿的每一个公式结果都与指标引擎一致")
    print()
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
